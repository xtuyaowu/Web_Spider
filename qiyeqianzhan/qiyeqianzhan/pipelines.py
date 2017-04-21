# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import xlsxwriter
from openpyxl import load_workbook

class QiyeqianzhanPipeline(object):
    def __init__(self):
        workbook = xlsxwriter.Workbook(r'huizong.xlsx')
        worksheet = workbook.add_worksheet(r'pay_company')
        workbook.close()

    def process_item(self, item, spider):
        # print(item['compony_name'].strip(), "抓取成功")
        wb = load_workbook(r'huizong.xlsx')
        ws = wb.get_sheet_by_name(r'pay_company')
        col1 = 'A' + item['id']
        col2 = 'B' + item['id']
        col3 = 'C' + item['id']
        col4 = 'D' + item['id']
        col5 = 'E' + item['id']
        col6 = 'F' + item['id']
        col7 = 'G' + item['id']
        col8 = 'H' + item['id']
        col9 = 'I' + item['id']
        col10 = 'J' + item['id']
        col11 = 'K' + item['id']
        col12 = 'L' + item['id']
        col13 = 'M' + item['id']
        col14 = 'N' + item['id']
        col15 = 'O' + item['id']
        col16 = 'P' + item['id']
        col17 = 'Q' + item['id']
        col18 = 'R' + item['id']
        col19 = 'S' + item['id']
        col20 = 'T' + item['id']
        ws[col1] = item['compony_name']
        ws[col2] = item['gsinfo_shxydm']
        ws[col3] = item['gsinfo_zch']
        ws[col4] = item['gsinfo_jgdm']
        ws[col5] = item['gsinfo_jyzt']
        ws[col6] = item['gsinfo_ztlx']
        ws[col7] = item['gsinfo_jglx']
        ws[col8] = item['gsinfo_djrq']
        ws[col9] = item['gsinfo_frdb']
        ws[col10] = item['gsinfo_zczb']
        ws[col11] = item['gsinfo_jyxq']
        ws[col12] = item['gsinfo_djjg']
        ws[col13] = item['gsinfo_ssdq']
        ws[col14] = item['gsinfo_zsyxq']
        ws[col15] = item['gsinfo_fzrq']
        ws[col16] = item['gsinfo_qzbq']
        ws[col17] = item['gsinfo_zhbq']
        ws[col18] = item['gsinfo_sshy']
        ws[col19] = item['compony_adress']
        ws[col20] = item['gsinfo_jyfw']
        wb.save(r'huizong.xlsx')
        print(item,"写入成功")

        file = open(r'companylist.txt', mode='a',
                    encoding='utf8')
        if item['compony_name']:
            file.writelines(item['compony_name']+"\n")
        else:
            file.writelines("抓取失败")
        if item['compony_url']:
            file.writelines(item['compony_url']+"\n")
        else:
            file.writelines("抓取失败")
        file.close()

        return item
